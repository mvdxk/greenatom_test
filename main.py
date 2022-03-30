import requests
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Alignment
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
import os
import mimetypes
from email import encoders

moex_url = 'https://www.moex.com/ru/derivatives/currency-rate.aspx?currency='
currs = {
    'usd': 'USD_RUB',
    'eur': 'EUR_RUB'
}
res = {
    'usd': '',
    'eur': ''
}
thead = ['Дата', 'Значение курса промежуточного клиринга ', 'Время', 'Значение курса основного клиринга ', 'Время']

usd = []
eur = []

def makeDict(vals):
    dict = []
    for str in vals:
        a = []
        for s in str.find_all('td'):
            a.append((s.string + '').replace(',', '.'))
        dict.append(a)
    return dict

def join(usd, eur):
    data = {}
    el = ['-', '-', '-', '-', '-']
    for u in usd:
        data[u[0]] = u + el
    for e in eur:
        if list(data.keys()).count(e[0]) == 0:
            data[e[0]] = el + e[1:]
        else:
            data[e[0]] = data[e[0]][:5] + e
    return data

def createFile(data):
    file = xl.Workbook()
    sheet = file.active
    sheet['J1'] = 'Изменение'
    sheet.cell(row=1, column=9).alignment = Alignment(horizontal='center')
    col = 1
    for c in range(len(currs)):
        for t in range(len(thead)):
            h = thead[t]
            if h.find('курса') != -1:
                h += currs[list(currs.keys())[c]]
            sheet.cell(row=1, column=col, value=h)
            sheet.cell(row=1, column=col).alignment = Alignment(horizontal='center')
            col += 1
    row = 2

    for k in data.keys():
        d = data[k]
        for i in range(9):
            val = d[i]
            if val.count(':') == 0 and val.count('.') == 1:
                val = float(val)
            sheet.cell(row=row, column=i + 1, value=val)
            sheet.cell(row=row, column=i + 1).alignment = Alignment(horizontal='center')

        val = '-'
        if d[3] != '-' and d[-2] != '-':
            val = float(d[-2]) / float(d[3])
        sheet.cell(row=row, column=10, value=val)
        sheet.cell(row=row, column=10).alignment = Alignment(horizontal='center')
        row += 1

    for l in ['B', 'D', 'G',  'I']:
        sheet.column_dimensions[l].width = 46
    file.save('data.xlsx')

def sendEmail(strAmount):
    addr_from = "test.greenatom@mail.ru"
    addr_to = "test.greenatom@mail.ru"
    password = "BZCi8XdGKj4YVqDFFiad"

    msg = MIMEMultipart()
    msg['From'] = addr_from
    msg['To'] = addr_to
    msg['Subject'] = '-'

    word = 'строка'
    if str(strAmount)[-1] == '0' or 4 < strAmount < 20:
        word = 'строк'
    elif str(strAmount)[-1] != 1 and 1 < strAmount%10 < 5:
        word = 'строки'

    body = str(strAmount) + ' ' + word
    msg.attach(MIMEText(body, 'plain'))

    filepath = 'C:\\Users\\Sas\\Documents\\GitHub\\greenatom_test\\data.xlsx'
    filename = os.path.basename(filepath)
    ctype, encoding = mimetypes.guess_type(filepath)
    maintype, subtype = ctype.split('/', 1)
    if os.path.isfile(filepath):
        with open(filepath, 'rb') as fp:
            file = MIMEBase(maintype, subtype)
            file.set_payload(fp.read())
            fp.close()
        encoders.encode_base64(file)
    file.add_header('Content-Disposition', 'attachment', filename=filename)
    msg.attach(file)

    server = smtplib.SMTP_SSL('smtp.mail.ru', 465)
    server.login(addr_from, password)
    server.send_message(msg)
    server.quit()




res['usd'] = BeautifulSoup(requests.get(moex_url + currs['usd']).text, 'lxml')
res['eur'] = BeautifulSoup(requests.get(moex_url + currs['eur']).text, 'lxml')
res['usd'] = res['usd'].find('table', class_='tablels').find_all('tr')[2:]
res['eur'] = res['eur'].find('table', class_='tablels').find_all('tr')[2:]
usd = makeDict(res['usd'])
eur = makeDict(res['eur'])
data = join(usd, eur)
createFile(data)
sendEmail(len(data)+1)
